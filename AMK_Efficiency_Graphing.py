import csv
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

Filename = "/Users/srijitmukherjee/Documents/University_Subjects/UQ Racing/Inverters and BMS Research/Motor Efficiency Data Spreadsheet.xlsx"

#Reading the excel file and storing it as a workbook object
wb_obj = openpyxl.load_workbook(Filename)

#Selecting the efficiency sheet
eff_sheet = wb_obj["Efficiency Sheet"]
speed_sheet = wb_obj["Speed Sheet"]
torque_sheet = wb_obj["Torque Sheet"]


#Initialise Arrays
eff_array = np.zeros((100,100), dtype = float)
speed_array = np.zeros((100,100), dtype = float)
torque_array = np.zeros((100, 100), dtype = float)

#Now store excel values in the array

def populate_array(array, sheet = str):
    for index, element in np.ndenumerate(array):
        #Here the index will be a 2d row, col vector
        #The element will be what's contained in the array
        if sheet == "Speed Sheet":
            
            #Getting the value of the cell and then storing it in the array
            sheet_cell_val = speed_sheet.cell(row = index[0] + 1, column = index[1] +1).value
            array[index[0]][index[1]] = sheet_cell_val

        elif sheet == "Torque Sheet":
            sheet_cell_val = torque_sheet.cell(row = index[0] + 1, column = index[1] +1).value
            array[index[0]][index[1]] = sheet_cell_val

        else:
            sheet_cell_val = eff_sheet.cell(row = index[0] + 1, column = index[1] +1).value
            array[index[0]][index[1]] = sheet_cell_val


populate_array(eff_array, "Efficiency Sheet")
populate_array(speed_array, "Speed Sheet")
populate_array(torque_array, "Torque Sheet")


#Setting up the contour levels and removing the Nan values
levels = np.linspace(0.75, np.nanmax(eff_array), 15)
fig, ax = plt.subplots()
CS = ax.contourf(torque_array, speed_array, eff_array, cmap = 'rainbow', levels=levels)
plt.xlabel("Torque")
plt.ylabel("Speed")
plt.title("The effect of motor speed and torque on efficiency")
cbar = fig.colorbar(CS)
plt.show()






            
            
            

            
        
        


